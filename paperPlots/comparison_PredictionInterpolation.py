import sys
sys.path.append(".")
import numpy as np
import pandas as pd
from create_data import callFunction
from plot import predict_encoder_interpolation_projection_roundedInput
from useful import remove_parts_of_graph_encoder
from numpy import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from scipy.interpolate import PchipInterpolator
import statsmodels.api as sm
from scipy.signal import savgol_filter

def one_runs_from_mask(mask_01:pd.DataFrame):
    """
    mask_01: 1D-Array/List/Series mit 0/1 (oder bool)
    Rückgabe: Liste von (start_idx, end_idx) inklusiv
    """
    s = pd.Series(mask_01).astype(bool)  # True=1, False=0

    # Startpunkte: Stelle ist True und davor False/Anfang
    starts = s & ~s.shift(fill_value=False)
    # Endpunkte: Stelle ist True und danach False/Ende
    ends   = s & ~s.shift(-1, fill_value=False)

    start_idx = np.flatnonzero(starts.to_numpy())
    end_idx   = np.flatnonzero(ends.to_numpy())
    return list(zip(start_idx, end_idx))



def saveToWorksheet(ws: Worksheet, col_letter:str, col_id:int, timeSeries):
    ws[f"{col_letter}1"] = f"TimeSeries_{col_id}"
    # Daten ab Zeile 2 runter schreiben
    for row_idx, value in enumerate(timeSeries, start=2):
        ws[f"{col_letter}{row_idx}"].value = float(value)



def denoiseTimeSeries(model_input: str, mask):
    prediction_encoder = predict_encoder_interpolation_projection_roundedInput(len(x_values), model_input, y_trend_noise, min_value, max_value, vocabSize, extraToken, mask)
    prediction_encoder = prediction_encoder.squeeze()
    prediction_encoder = prediction_encoder.numpy()
    return prediction_encoder

def calcMaximumAbsoluteError(prediction, groundTruth):
    error = (prediction - groundTruth) / noise_std_value
    error = abs(error)
    error = max(error)
    return error

def calcMeanAbsoluteError(prediction, groundTruth):
    error = (prediction - groundTruth) / noise_std_value
    error = abs(error)
    error = np.mean(error)
    return error



def smooth_segments(s: pd.Series,
                    method: str = "savgol",
                    window: int = 7,      # Anzahl Punkte (ungerade) für SavGol/rolling
                    polyorder: int = 2,   # SavGol Polynomgrad
                    loess_frac: float = 0.2,
                    ewma_alpha: float = 0.2):
    """
    Glättet NUR innerhalb zusammenhängender nicht-NaN-Segmente.
    Gibt eine Serie mit NaNs an den ursprünglichen Lücken zurück.
    """
    s = s.copy()
    isna = s.isna()
    groups = (isna.ne(isna.shift())).cumsum()  # Segment-IDs
    out = pd.Series(index=s.index, dtype=float)

    for gid, seg in s.groupby(groups):
        if seg.isna().all():  # Das ist eine NaN-Lücke -> überspringen
            continue
        y = seg.values.astype(float)

        if method == "rolling_mean":
            y_hat = pd.Series(y, index=seg.index).rolling(window, min_periods=1, center=True).mean().values
        elif method == "rolling_median":
            y_hat = pd.Series(y, index=seg.index).rolling(window, min_periods=1, center=True).median().values
        elif method == "ewma":
            y_hat = pd.Series(y, index=seg.index).ewm(alpha=ewma_alpha, adjust=False).mean().values
        elif method == "loess":
            x = np.arange(len(y))
            y_hat = sm.nonparametric.lowess(y, x, frac=loess_frac, return_sorted=False)
        elif method == "savgol":
            win = window if window % 2 == 1 else window + 1
            if len(y) >= win:
                y_hat = savgol_filter(y, window_length=win, polyorder=min(polyorder, win-1), mode="interp")
            else:
                # Fallback: kleine Segmente -> Median-Rolling
                y_hat = pd.Series(y, index=seg.index).rolling(min(len(y), 3), min_periods=1, center=True).median().values
        elif method == "none":
            y_hat = y
        else:
            raise ValueError("Unbekannte Glättungsmethode")

        out.loc[seg.index] = y_hat

    return out

def interpolate_gaps(s: pd.Series,
                     method: str = "pchip",
                     max_gap: int | None = None,
                     datetime_aware: bool = True):
    """
    Interpoliert NUR NaN-Lücken. Optional: nur Lücken bis max_gap Länge.
    method: 'time' (bei DatetimeIndex), 'pchip', 'spline3', 'linear', 'polynomial3', 'kalman'
    """
    s = s.copy()

    # Optional: nur kleine/mittlere Lücken interpolieren
    if max_gap is not None:
        isna = s.isna()
        grp = (isna.ne(isna.shift())).cumsum()
        gap_lens = isna.groupby(grp).transform('sum')
        mask_large_gaps = isna & (gap_lens > max_gap)
    else:
        mask_large_gaps = pd.Series(False, index=s.index)

    if method == "time":
        if not isinstance(s.index, (pd.DatetimeIndex, pd.PeriodIndex)):
            raise ValueError("method='time' benötigt DatetimeIndex.")
        filled = s.interpolate(method="time", limit_area="inside")
    elif method == "linear":
        filled = s.interpolate(method="linear", limit_area="inside")
    elif method == "pchip":
        # PCHIP via SciPy
        x = np.arange(len(s))
        ok = ~s.isna()
        f = PchipInterpolator(x[ok], s[ok])
        filled = pd.Series(f(x), index=s.index)
        # Ränder (Extrapolation) auf Original lassen, falls dort NaN:
        filled[~ok & ((~ok).cummax() & (~ok)[::-1].cummax()[::-1])] = np.nan
    elif method == "spline3":
        filled = s.interpolate(method="spline", order=3, limit_area="inside")
    elif method == "polynomial3":
        filled = s.interpolate(method="polynomial", order=3, limit_area="inside")
    elif method == "kalman":
        # sehr einfache Zustandsraum-Variante (lokaler Trend)
        mod = sm.tsa.UnobservedComponents(s, level='local linear trend')
        res = mod.fit(disp=False)
        filled = pd.Series(res.predict(), index=s.index)
    else:
        raise ValueError("Unbekannte Interpolationsmethode")

    # Große Lücken bewusst offen lassen
    filled[mask_large_gaps] = np.nan
    return filled

def smooth_then_interpolate(s: pd.Series,
                            smooth_method="savgol",
                            interp_method="pchip",
                            smooth_window=7,
                            max_gap=None):
    """
    1) innerhalb sichtbarer Segmente glätten
    2) NaN-Lücken interpolieren (kontrolliert)
    """
    s_sm = smooth_segments(s, method=smooth_method, window=smooth_window)
    out = interpolate_gaps(s_sm, method=interp_method, max_gap=max_gap)
    return out








if __name__ == "__main__":
    '''

    Todo:
    -arrays erstellen, um die Fehler (maximaler absolute Error) zwischenzuspeichern
        -error_LowOrder
        -error_Periodic
        -error_PeriodicSum
        -error_Discontinuous
        -error_AllInOne
        -error_Wavelet (wavelet="db6", level=None, threshold_rule="universal", threshold_scale=1.0, mode="soft", signal_extension="symmetric", use_swt=False)
        -error_Fourier (sampling rate: 500Hz, cutoff Frequency: 10Hz)
    -dataframe erstellen, in welchem die Fehler eingetragen werden
    for timeSeriesOption in timeSeriesGenerator:
        for schleife mit der anzahl 100:
        
            -Zeitreihe synthetisch erstellen (timeseries choose aus erster for schleife variable nehmen)
            -noisy data glätten 
                Low Order
                Discontinuous
                Periodic
                Periodic Sum
                All In One
                Wavelet
                Fourier 
            -fehler (maximum absolute error) in arrays einpflegen
    
    -fehler durchschnitt berechnen
    -fehler in ein dataframe packen (jede Spalte den jewiligen Funktionstyp (low order, high order, periodic, periodic sum, discontinuous))
    
    '''

    error_MeanAbsoluteMax_lowOrder = []
    error_MeanAbsoluteMax_Periodic = []
    error_MeanAbsoluteMax_PeriodicSum = []
    error_MeanAbsoluteMax_AllInOne = []
    error_MeanAbsoluteMax_LinearInterpolation = []
    error_MeanAbsoluteMax_PolynomialInterpolation = []

    error_MeanAbsolute_lowOrder = []
    error_MeanAbsolute_Periodic = []
    error_MeanAbsolute_PeriodicSum = []
    error_MeanAbsolute_AllInOne = []
    error_MeanAbsolute_LinearInterpolation = []
    error_MeanAbsolute_PolynomialInterpolation = []


    dataframe_MeanAbsoluteMax_Error = pd.DataFrame(columns=["LowOrder", "Periodic", "PeriodicSum", "AllInOne"], index=["ModelLowOrder", "Model Periodic", "Model Periodic Sum", "Model All In One", "Linear Interpolation", "Polynomial3 Interpolation"])
    dataframe_MeanAbsolute_Error = pd.DataFrame(columns=["LowOrder", "Periodic", "PeriodicSum", "AllInOne"], index=["ModelLowOrder", "Model Periodic", "Model Periodic Sum", "Model All In One", "Linear Interpolation", "Polynomial3 Interpolation"])

    modelName_timeseries = {
    "LowOrder": [0,1,2],
    "Periodic":[3],
    "PeriodicSum": [4],
    "AllInOne": [0,1,2,3,4],
    }


    #settings
    x_values = np.arange(1000)
    y_start = random.uniform(-10,10)
    random_number_range = ["norm",0,5]              # ["uni", lowerRange, upperRange], ["norm", mean, std]
    splineValue = [800000,1100000]
    vocabSize = 100000
    extraToken = ["ukn"]                            # UKN Token
    noise_std = ["norm",0,0.1]                     # ["uni", lowerRange, upperRange], ["norm", mean, std]




    for i, (k,v) in enumerate(modelName_timeseries.items()):

        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="Groundtruth") #0
        wb.create_sheet(title="Noisy_Data") #1
        wb.create_sheet(title="Low_Order") #2
        wb.create_sheet(title="Periodic") #3
        wb.create_sheet(title="Periodic_Sum") #4
        wb.create_sheet(title="All_In_One") #5
        wb.create_sheet(title="Linear_Interpolation")#6
        wb.create_sheet(title="Polynomial_Interpolation") #7
        wb.create_sheet(title="mask") #8


        if k == "LowOrder":
            splineValue = [800000,1100000]
        elif k == "AllInOne":
            splineValue = [100000,1100000]
        else: 
            None
        
        for j in range(1,101):
            col_letter = get_column_letter(j)

            # fig, ax = plt.subplots(1,1)
            print(f"Iteration: {j}")
            timeSeriesChoose = np.random.choice(v)

            y_trend, y_trend_noise, min_value, max_value, noise_std_value = callFunction(x_values, y_start, random_number_range, splineValue, vocabSize, timeSeriesChoose, noise_std)
            # ax.plot(x_values, y_trend_noise, label="Noise")
            # ax.plot(x_values, y_trend, label="GroundTruth", linewidth = 5)
            saveToWorksheet(wb.worksheets[0], col_letter, j, y_trend)
            saveToWorksheet(wb.worksheets[1], col_letter, j, y_trend_noise)

            mask = remove_parts_of_graph_encoder(np.arange(1000), y_trend_noise,[10,100,10],10, [0,1000])
            saveToWorksheet(wb.worksheets[8], col_letter, j, mask)



            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_LowOrder_300.pt", mask)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_lowOrder.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_lowOrder.append(error)
            # ax.plot(x_values, prediction, label="Low Order")
            saveToWorksheet(wb.worksheets[2], col_letter, j, prediction)

            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_Periodic_300.pt", mask)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_Periodic.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_Periodic.append(error)
            # ax.plot(x_values, prediction, label="Periodic")
            saveToWorksheet(wb.worksheets[3], col_letter, j, prediction)

            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_PeriodicSum_2000.pt", mask)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_PeriodicSum.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_PeriodicSum.append(error)
            # ax.plot(x_values, prediction, label="Periodic Sum")
            saveToWorksheet(wb.worksheets[4], col_letter, j, prediction)

            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_AllInOne_2400.pt", mask)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_AllInOne.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_AllInOne.append(error)
            # ax.plot(x_values, prediction, label="All In One")
            saveToWorksheet(wb.worksheets[5], col_letter, j, prediction)

            y_trend_noise[mask == 1] = np.nan
            y_trend_noise = pd.Series(y_trend_noise)

            print("Preloading Linear Interpolation")
            prediction = smooth_then_interpolate(y_trend_noise,
                                        smooth_method="loess",   # oder 'loess', 'rolling_median', 'ewma', 'none', savgol
                                        interp_method="linear",    # oder 'time','spline3','linear','polynomial3','kalman', 'pchip
                                        smooth_window=9,
                                        max_gap=1000)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_LinearInterpolation.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_LinearInterpolation.append(error)
            # ax.plot(x_values, prediction, label="All In One")
            saveToWorksheet(wb.worksheets[6], col_letter, j, prediction)
        

            print("Preloading Polynomial3 Interpolation")
            prediction = smooth_then_interpolate(y_trend_noise,
                                        smooth_method="loess",   # oder 'loess', 'rolling_median', 'ewma', 'none', savgol
                                        interp_method="polynomial3",    # oder 'time','spline3','linear','polynomial3','kalman', 'pchip
                                        smooth_window=9,
                                        max_gap=1000)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_PolynomialInterpolation.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_PolynomialInterpolation.append(error)
            # ax.plot(x_values, prediction, label="All In One")
            saveToWorksheet(wb.worksheets[7], col_letter, j, prediction)

            print("\n")

            # plt.legend()
            # plt.show()

            if j % 20 == 0:
                print("Mean Absolute Max Error:")
                print(f"Low Order: {np.mean(error_MeanAbsoluteMax_lowOrder)},\nPeriodic: {np.mean(error_MeanAbsoluteMax_Periodic)},\nPeriodicSum: {np.mean(error_MeanAbsoluteMax_PeriodicSum)},\nAllInOne {np.mean(error_MeanAbsoluteMax_AllInOne)},\nLinear Interpolation: {np.mean(error_MeanAbsoluteMax_LinearInterpolation)},\nPolynomial3 Interpolation: {np.mean(error_MeanAbsoluteMax_PolynomialInterpolation)}")
                print("\n")
                print("Mean Absolute Error:")
                print(f"Low Order: {np.mean(error_MeanAbsolute_lowOrder)},\nPeriodic: {np.mean(error_MeanAbsolute_Periodic)},\nPeriodicSum: {np.mean(error_MeanAbsolute_PeriodicSum)},\nAllInOne {np.mean(error_MeanAbsolute_AllInOne)},\nLinear Interpolation: {np.mean(error_MeanAbsolute_LinearInterpolation)},\nPolynomial3 Interpolation: {np.mean(error_MeanAbsolute_PolynomialInterpolation)}")
                print("\n")


        dataframe_MeanAbsoluteMax_Error.loc[:,k] = [np.mean(error_MeanAbsoluteMax_lowOrder), np.mean(error_MeanAbsoluteMax_Periodic), np.mean(error_MeanAbsoluteMax_PeriodicSum), np.mean(error_MeanAbsoluteMax_AllInOne), np.mean(error_MeanAbsoluteMax_LinearInterpolation), np.mean(error_MeanAbsoluteMax_PolynomialInterpolation)]
        dataframe_MeanAbsolute_Error.loc[:,k] = [np.mean(error_MeanAbsolute_lowOrder), np.mean(error_MeanAbsolute_Periodic), np.mean(error_MeanAbsolute_PeriodicSum), np.mean(error_MeanAbsolute_AllInOne), np.mean(error_MeanAbsolute_LinearInterpolation),np.mean(error_MeanAbsolute_PolynomialInterpolation)]

        print("Mean Absolute Max Error:")
        print(dataframe_MeanAbsoluteMax_Error)
        print("\n")
        print("Mean Absolute Error:")
        print(dataframe_MeanAbsolute_Error)
        print("\n")
        
        error_MeanAbsoluteMax_lowOrder = []
        error_MeanAbsoluteMax_Periodic = []
        error_MeanAbsoluteMax_PeriodicSum = []
        error_MeanAbsoluteMax_AllInOne = []
        error_MeanAbsoluteMax_LinearInterpolation = []
        error_MeanAbsoluteMax_PolynomialInterpolation = []

        error_MeanAbsolute_lowOrder = []
        error_MeanAbsolute_Periodic = []
        error_MeanAbsolute_PeriodicSum = []
        error_MeanAbsolute_AllInOne = []
        error_MeanAbsolute_LinearInterpolation = []
        error_MeanAbsolute_PolynomialInterpolation = []

        wb.save(f"data/error_comparison_interpolation/{k}.xlsx")
    
    dataframe_MeanAbsoluteMax_Error.to_csv("data/error_comparison_interpolation/Mean_Absolute_Max_Error.csv", index=True)
    dataframe_MeanAbsolute_Error.to_csv("data/error_comparison_interpolation/Mean_Absolute_Error.csv", index=True)

            





