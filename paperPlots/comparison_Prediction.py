import sys
sys.path.append(".")
import numpy as np
import pandas as pd
from create_data import callFunction
from plot import predict_encoder_interpolation_projection_roundedInput
from useful import sliding_window
from numpy import random
import pywt
from scipy.signal import butter, filtfilt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def saveToWorksheet(ws: Worksheet, col_letter:str, col_id:int, timeSeries):
    ws[f"{col_letter}1"] = f"TimeSeries_{col_id}"
    # Daten ab Zeile 2 runter schreiben
    for row_idx, value in enumerate(timeSeries, start=2):
        ws[f"{col_letter}{row_idx}"].value = float(value)



def denoiseTimeSeries(model_input: str):
    prediction_encoder = predict_encoder_interpolation_projection_roundedInput(len(x_values), model_input, y_trend_noise, min_value, max_value, vocabSize, extraToken, np.zeros_like(x_values))
    prediction_encoder = prediction_encoder.squeeze()
    prediction_encoder = prediction_encoder.numpy()
    return prediction_encoder

def calcMaximumAbsoluteError(prediction, groundTruth):
    error = (prediction - groundTruth) / noise_std_value
    error = abs(error)
    error = max(error)
    return error

def calcMeanAbsoluteError(prediction, groundTruth):
    error = (prediction - groundTruth) / noise_std_value
    error = abs(error)
    error = np.mean(error)
    return error

def butter_lowpass(cutoff, fs, order=4):
        nyq = 0.5 * fs
        normal_cutoff = cutoff / nyq
        b, a = butter(order, normal_cutoff, btype='low', analog=False)
        return b, a

def lowpass_filter(data, cutoff, fs, order=4):
    b, a = butter_lowpass(cutoff, fs, order=order)
    y = filtfilt(b, a, data)
    return y

def wavelet_denoise(
        x,
        wavelet="db4",          # Wavelet-Familie / -Ordnung (z.B. "db4", "sym8", "coif5", "bior4.4")
        level=None,             # Zerlegungstiefe; None = automatisch (max. sinnvoll)
        threshold_rule="universal",  # 'universal' (VisuShrink), 'sure' (SURE), 'bayes' (BayesShrink)
        threshold_scale=1.0,    # Faktor für die Schwelle (z.B. 0.8..1.2 Feintuning)
        mode="soft",            # 'soft' oder 'hard' Thresholding
        signal_extension="symmetric",# Randbehandlung: 'symmetric','periodization','reflect','zero', ...
        use_swt=False           # True: Stationary Wavelet Transform (übersetzungsinvariant)
    ):
        """
        Gibt das geglättete Signal zurück und ein Dict mit Metadaten.
        """
        # Hilfsfunktionen
        def noise_sigma_from_coeffs(detail_coeffs):
            # Rauschschätzer (MAD) aus feinster Detail-Ebene
            d1 = detail_coeffs[-1]
            sigma = np.median(np.abs(d1 - np.median(d1))) / 0.6745
            return sigma

        def calc_threshold(sigma, n, rule):
            if rule == "universal":         # VisuShrink
                return sigma * np.sqrt(2*np.log(n))
            elif rule == "sure":            # Stein's Unbiased Risk Estimate (PyWavelets hat Hilfen, hier simple Approx.)
                # Vereinfachte SURE-Variante: starte bei universal und skaliere
                return sigma * np.sqrt(2*np.log(n)) * 0.8
            elif rule == "bayes":
                # BayesShrink (vereinfachte Heuristik): sigma^2 / std(detail)
                return None  # pro Ebene separat weiter unten
            else:
                raise ValueError("threshold_rule must be 'universal', 'sure', or 'bayes'.")

        # Zerlegung
        if use_swt:
            coeffs = pywt.swt(x, wavelet=wavelet, level=level or pywt.swt_max_level(len(x)), start_level=0, trim_approx=True)
            # coeffs: Liste [(cA_L, cD_L), ..., (cA_1, cD_1)]
            sigma = noise_sigma_from_coeffs([cD for (_, cD) in coeffs])
            n = len(x)
            T_global = calc_threshold(sigma, n, threshold_rule)
            denoised_coeffs = []
            for (cA, cD) in coeffs:
                if threshold_rule == "bayes":
                    sigma_d = np.std(cD)
                    T = (sigma**2) / (sigma_d + 1e-12)
                else:
                    T = T_global
                T = threshold_scale * T if T is not None else None
                cD_th = pywt.threshold(cD, T, mode=mode) if T is not None else cD
                denoised_coeffs.append((cA, cD_th))
            x_hat = pywt.iswt(denoised_coeffs, wavelet=wavelet)
        else:
            max_level = pywt.dwt_max_level(data_len=len(x), filter_len=pywt.Wavelet(wavelet).dec_len)
            use_level = min(level or max_level, max_level)
            coeffs = pywt.wavedec(x, wavelet=wavelet, mode=signal_extension, level=use_level)
            cA, cDs = coeffs[0], coeffs[1:]
            sigma = noise_sigma_from_coeffs(cDs)
            n = len(x)
            T_global = calc_threshold(sigma, n, threshold_rule)

            new_cDs = []
            for cD in cDs:
                if threshold_rule == "bayes":
                    sigma_d = np.std(cD)
                    T = (sigma**2) / (sigma_d + 1e-12)
                else:
                    T = T_global
                T = threshold_scale * T if T is not None else None
                cD_th = pywt.threshold(cD, T, mode=mode) if T is not None else cD
                new_cDs.append(cD_th)

            x_hat = pywt.waverec([cA] + new_cDs, wavelet=wavelet, mode=signal_extension)

        return x_hat[:len(x)], {
            "wavelet": wavelet,
            "level": level,
            "threshold_rule": threshold_rule,
            "threshold_scale": threshold_scale,
            "mode": mode,
            "signal_extension": signal_extension,
            "use_swt": use_swt
        }





if __name__ == "__main__":
    '''

    Todo:
    -arrays erstellen, um die Fehler (maximaler absolute Error) zwischenzuspeichern
        -error_LowOrder
        -error_Periodic
        -error_PeriodicSum
        -error_Discontinuous
        -error_AllInOne
        -error_Wavelet (wavelet="db6", level=None, threshold_rule="universal", threshold_scale=1.0, mode="soft", signal_extension="symmetric", use_swt=False)
        -error_Fourier (sampling rate: 500Hz, cutoff Frequency: 10Hz)
    -dataframe erstellen, in welchem die Fehler eingetragen werden
    for timeSeriesOption in timeSeriesGenerator:
        for schleife mit der anzahl 100:
        
            -Zeitreihe synthetisch erstellen (timeseries choose aus erster for schleife variable nehmen)
            -noisy data glätten 
                Low Order
                Discontinuous
                Periodic
                Periodic Sum
                All In One
                Wavelet
                Fourier 
            -fehler (maximum absolute error) in arrays einpflegen
    
    -fehler durchschnitt berechnen
    -fehler in ein dataframe packen (jede Spalte den jewiligen Funktionstyp (low order, high order, periodic, periodic sum, discontinuous))
    
    '''

    error_MeanAbsoluteMax_lowOrder = []
    error_MeanAbsoluteMax_Periodic = []
    error_MeanAbsoluteMax_PeriodicSum = []
    error_MeanAbsoluteMax_AllInOne = []
    error_MeanAbsoluteMax_Wavelet = []
    error_MeanAbsoluteMax_Fourier = []
    error_MeanAbsoluteMax_SlidingWindow = []

    error_MeanAbsolute_lowOrder = []
    error_MeanAbsolute_Periodic = []
    error_MeanAbsolute_PeriodicSum = []
    error_MeanAbsolute_AllInOne = []
    error_MeanAbsolute_Wavelet = []
    error_MeanAbsolute_Fourier = []
    error_MeanAbsolute_SlidingWindow = []


    dataframe_MeanAbsoluteMax_Error = pd.DataFrame(columns=["LowOrder", "Periodic", "PeriodicSum", "AllInOne"], index=["ModelLowOrder", "Model Periodic", "Model Periodic Sum", "Model All In One", "Wavelet", "Fourier", "Sliding_Window"])
    dataframe_MeanAbsolute_Error = pd.DataFrame(columns=["LowOrder", "Periodic", "PeriodicSum", "AllInOne"], index=["ModelLowOrder", "Model Periodic", "Model Periodic Sum", "Model All In One", "Wavelet", "Fourier","Sliding_Window"])

    modelName_timeseries = {
    "LowOrder": [0,1,2],
    "Periodic":[3],
    "PeriodicSum": [4],
    "AllInOne": [0,1,2,3,4],
    }


    #settings
    x_values = np.arange(1000)
    y_start = random.uniform(-10,10)
    random_number_range = ["norm",0,5]              # ["uni", lowerRange, upperRange], ["norm", mean, std]
    splineValue = [800000,1100000]
    vocabSize = 100000
    extraToken = ["ukn"]                            # UKN Token
    noise_std = ["norm",0,0.1]                     # ["uni", lowerRange, upperRange], ["norm", mean, std]




    for i, (k,v) in enumerate(modelName_timeseries.items()):

        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="Groundtruth") #0
        wb.create_sheet(title="Noisy_Data") #1
        wb.create_sheet(title="Low_Order") #2
        wb.create_sheet(title="Periodic") #3
        wb.create_sheet(title="Periodic_Sum") #4
        wb.create_sheet(title="All_In_One") #5
        wb.create_sheet(title="Wavelet") #6
        wb.create_sheet(title="Fourier") #7
        wb.create_sheet(title="Sliding_Window") #8


        if k == "LowOrder":
            splineValue = [800000,1100000]
        elif k == "AllInOne":
            splineValue = [100000,1100000]
        else: 
            None
        
        for j in range(1,101):
            col_letter = get_column_letter(j)

            # fig, ax = plt.subplots(1,1)
            print(f"Iteration: {j}")
            timeSeriesChoose = np.random.choice(v)

            y_trend, y_trend_noise, min_value, max_value, noise_std_value = callFunction(x_values, y_start, random_number_range, splineValue, vocabSize, timeSeriesChoose, noise_std)
            # ax.plot(x_values, y_trend_noise, label="Noise")
            # ax.plot(x_values, y_trend, label="GroundTruth", linewidth = 5)
            saveToWorksheet(wb.worksheets[0], col_letter, j, y_trend)
            saveToWorksheet(wb.worksheets[1], col_letter, j, y_trend_noise)


            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_LowOrder_300.pt")
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_lowOrder.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_lowOrder.append(error)
            # ax.plot(x_values, prediction, label="Low Order")
            saveToWorksheet(wb.worksheets[2], col_letter, j, prediction)

            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_Periodic_300.pt")
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_Periodic.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_Periodic.append(error)
            # ax.plot(x_values, prediction, label="Periodic")
            saveToWorksheet(wb.worksheets[3], col_letter, j, prediction)

            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_PeriodicSum_2000.pt")
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_PeriodicSum.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_PeriodicSum.append(error)
            # ax.plot(x_values, prediction, label="Periodic Sum")
            saveToWorksheet(wb.worksheets[4], col_letter, j, prediction)

            prediction = denoiseTimeSeries(r"C:\Users\larsr\Documents\Uni\Paper\Modelle\Encoder_Interpolation_CE_AllInOne_2400.pt")
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_AllInOne.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_AllInOne.append(error)
            # ax.plot(x_values, prediction, label="All In One")
            saveToWorksheet(wb.worksheets[5], col_letter, j, prediction)

            print("Preloading Model Wavelet")
            prediction, _ = wavelet_denoise(
            y_trend_noise,
            wavelet="db6",
            level=None,            # automatisch
            threshold_rule="universal",
            threshold_scale=1.0,
            mode="soft",
            signal_extension="symmetric",
            use_swt=False          # True ausprobieren für SWT (meist glatter, aber langsamer)
            )
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_Wavelet.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_Wavelet.append(error)
            # ax.plot(x_values, prediction, label="Wavelet")
            saveToWorksheet(wb.worksheets[6], col_letter, j, prediction)

            print("Preloading Model Fourier")
            fs = 500  # Abtastrate [Hz]
            cutoff = 10  # Cutoff-Frequenz (Hz)
            prediction = lowpass_filter(y_trend_noise, cutoff, fs)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_Fourier.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_Fourier.append(error)
            # ax.plot(x_values, prediction, label="Fourier")
            saveToWorksheet(wb.worksheets[7], col_letter, j, prediction)


            print("Preloading Sliding Window")
            prediction = sliding_window(y_trend_noise, 20)
            for i in range(9):
                prediction = sliding_window(prediction, 20)
            error = calcMaximumAbsoluteError(prediction, y_trend)
            error_MeanAbsoluteMax_SlidingWindow.append(error)
            error = calcMeanAbsoluteError(prediction, y_trend)
            error_MeanAbsolute_SlidingWindow.append(error)
            # ax.plot(x_values, prediction, label="Sliding Window")
            saveToWorksheet(wb.worksheets[8], col_letter, j, prediction)

            print("\n")

            # plt.legend()
            # plt.show()

            if j % 20 == 0:
                print("Mean Absolute Max Error:")
                print(f"Low Order: {np.mean(error_MeanAbsoluteMax_lowOrder)},\nPeriodic: {np.mean(error_MeanAbsoluteMax_Periodic)},\nPeriodicSum: {np.mean(error_MeanAbsoluteMax_PeriodicSum)},\nAllInOne {np.mean(error_MeanAbsoluteMax_AllInOne)},\nWavelet: {np.mean(error_MeanAbsoluteMax_Wavelet)},\nFourier: {np.mean(error_MeanAbsoluteMax_Fourier)},\nSliding Window: {np.mean(error_MeanAbsoluteMax_SlidingWindow)}")
                print("\n")
                print("Mean Absolute Error:")
                print(f"Low Order: {np.mean(error_MeanAbsolute_lowOrder)},\nPeriodic: {np.mean(error_MeanAbsolute_Periodic)},\nPeriodicSum: {np.mean(error_MeanAbsolute_PeriodicSum)},\nAllInOne {np.mean(error_MeanAbsolute_AllInOne)},\nWavelet: {np.mean(error_MeanAbsolute_Wavelet)},\nFourier: {np.mean(error_MeanAbsolute_Fourier)},\nSliding Window: {np.mean(error_MeanAbsolute_SlidingWindow)}")
                print("\n")


        dataframe_MeanAbsoluteMax_Error.loc[:,k] = [np.mean(error_MeanAbsoluteMax_lowOrder), np.mean(error_MeanAbsoluteMax_Periodic), np.mean(error_MeanAbsoluteMax_PeriodicSum), np.mean(error_MeanAbsoluteMax_AllInOne), np.mean(error_MeanAbsoluteMax_Wavelet), np.mean(error_MeanAbsoluteMax_Fourier), np.mean(error_MeanAbsoluteMax_SlidingWindow)]
        dataframe_MeanAbsolute_Error.loc[:,k] = [np.mean(error_MeanAbsolute_lowOrder), np.mean(error_MeanAbsolute_Periodic), np.mean(error_MeanAbsolute_PeriodicSum), np.mean(error_MeanAbsolute_AllInOne), np.mean(error_MeanAbsolute_Wavelet), np.mean(error_MeanAbsolute_Fourier), np.mean(error_MeanAbsolute_SlidingWindow)]

        print("Mean Absolute Max Error:")
        print(dataframe_MeanAbsoluteMax_Error)
        print("\n")
        print("Mean Absolute Error:")
        print(dataframe_MeanAbsolute_Error)
        print("\n")
        
        error_MeanAbsoluteMax_lowOrder = []
        error_MeanAbsoluteMax_Periodic = []
        error_MeanAbsoluteMax_PeriodicSum = []
        error_MeanAbsoluteMax_AllInOne = []
        error_MeanAbsoluteMax_Wavelet = []
        error_MeanAbsoluteMax_Fourier = []
        error_MeanAbsoluteMax_SlidingWindow = []

        error_MeanAbsolute_lowOrder = []
        error_MeanAbsolute_Periodic = []
        error_MeanAbsolute_PeriodicSum = []
        error_MeanAbsolute_AllInOne = []
        error_MeanAbsolute_Wavelet = []
        error_MeanAbsolute_Fourier = []
        error_MeanAbsolute_SlidingWindow = []

        wb.save(f"data/error_comparison_prediction/{k}.xlsx")
    
    dataframe_MeanAbsoluteMax_Error.to_csv("data/error_comparison_prediction/Mean_Absolute_Max_Error.csv", index=True)
    dataframe_MeanAbsolute_Error.to_csv("data/error_comparison_prediction/Mean_Absolute_Error.csv", index=True)

            





